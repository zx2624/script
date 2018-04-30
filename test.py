from __future__ import division 
from openpyxl import Workbook
from openpyxl import load_workbook

n=input("you ji zu shuju :")
a=[0 for i in range(n)]
for i in range(0,n):
	a[i]=input("shuju%d   " % i )
begin=input("begin of zasanguang:   ")
end=input("end of zasanguang:    ")

for file_in in range(0,n):
	filename="%d-1.xlsx" % a[i]
	filename_dark="%dms.xlsx" % a[i]
	filename_za="%d-2.xlsx" % a[i]
	filename_result="%d-result.xlsx" % a[i]
	#print filename_dark
	wb=load_workbook(filename)
	ws=wb.active
	wb_dark=load_workbook(filename_dark)
	ws_dark=wb_dark.active
	wb_result=Workbook()
	ws_result=wb_result.active
	wb_za=load_workbook(filename_za)
	ws_za=wb_za.active
	
	#print ws.title
	index1=int(ws.cell(row=7,column=1).value)
	index2=index1+1
	count_1=0
	count_2=0
	aver_1=0
	aver_2=0
	result_index=0	
	count_za=0
	aver_za=0.0
	for i in range(begin,end+1):
		aver_za=(aver_za*count_za+ws_za.cell(row =i,column=2).value)
		count_za+=1
	print aver_za	
#print index1
#print index2
	for i in range(7,2055):
		lamda=int(ws.cell(row=i,column=1).value)
		ws.cell(row=i,column=2).value-=ws_dark.cell(row=i,column=2).value
	
		if lamda==index1:

			d=aver_1*count_1
		
			aver_1=(d+ws.cell(row=i,column=2).value)/(count_1+1)
		  
			count_1+=1
		if lamda==index2:
		
			aver_2=(aver_2*count_2+ws.cell(row=i,column=2).value)/(count_2+1)
			count_2+=1

		if lamda!=index1 and lamda!=index2:
			if count_1+count_2==0:continue
			result_index+=1
			ws_result.cell(row=result_index,column=1).value=lamda-2
			ws_result.cell(row=result_index,column=2).value=(aver_1*count_1+aver_2*count_2)/(count_1+count_2)

		 	aver_1=aver_2
			aver_2=ws.cell(row=i,column=2).value
			count_1=count_2
		
			count_2=1
			index1+=1
			index2+=1
	
	result_index+=1
	ws_result.cell(row=result_index,column=1).value=lamda-1
	ws_result.cell(row=result_index,column=2).value=(aver_1*count_1+aver_2*count_2)/(count_1+count_2)
	wb_result.save(filename = filename_result)
	


	
