from __future__ import division 
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import timedelta, datetime


def calcu_minutes(t1,t2):

	if t1.second==59:
		m1=t1.minute+1
	else:
		m1=t1.minute
	if t2.second==59:
		m2=t2.minute+1
	else:
		m2=t2.minute
	d=t1.day-t2.day
	h=t1.hour-t2.hour
	m=m1-m2
	return h*60+m+d*24*60

n=input("how manny?:  ")
a=[0 for i in range(n)]
for i in range(0,n):
	a[i]=raw_input("input%d  " % i)
	
print "this may take a while......"	
	                      
for i in range(0,n):	
	filename="%s.xlsx" % a[i]
	filename_result="%s-result.xlsx" % a[i]
	wb=load_workbook(filename)
	ws=wb.active
	wb_result=Workbook()
	ws_result=wb_result.active
	time1=ws['A2'].value
	print time1
	print "sldfkj;aldsffls;llj"
	sum=[0,0,0,0,0,0,0,0]
	for i in range(0,8):	
		sum[i]=ws.cell(row=2,column=i+2).value
	#	print sum
	count=1
	index=3;
	result_i=1
	# print calcu_minutes(time1, ws['A5'].value)
	# print ws['A5'].value.minute 
	# print ws['A5'].value.second
	while(ws['A%d' % index].value):
		if calcu_minutes(time1,ws['A%d' % index].value)==0:
			for i in range(0,8):
				sum[i]+=ws.cell(row=index,column=i+2).value
			#print sum
			count+=1
	# 		print count
		if calcu_minutes(time1,ws['A%d' % index].value)==1:
			if time1.second==59:
				ws_result['A%d' % result_i]=(time1+timedelta(minutes=1)).strftime('%Y-%m-%d %H:%M')
			else:
				ws_result['A%d' % result_i]=time1.strftime('%Y-%m-%d %H:%M')
			for i in range(0,8):
				ws_result.cell(row=result_i,column=i+2).value=sum[i]/count
			for i in range(0,8):
				sum[i]=ws.cell(row=index,column=i+2).value
			count=1
			time1=ws.cell(row=index,column=1).value
			result_i+=1
		if calcu_minutes(time1,ws['A%d' % index].value)>1:
	# 		print calcu_minutes(time1,ws['A%d' % index].value)
	# 		print time1
			if time1.second==59:
				ws_result['A%d' % result_i]=(time1+timedelta(minutes=1)).strftime('%Y-%m-%d %H:%M')
			else:
				ws_result['A%d' % result_i]=time1.strftime('%Y-%m-%d %H:%M')
			for i in range(0,8):
				ws_result.cell(row=result_i,column=i+2).value=sum[i]/count
			for i in range(0,8):
				sum[i]=ws.cell(row=index,column=i+2).value
			result_i+=1
			for i in range(1,calcu_minutes(time1,ws['A%d' % index].value)):
				if time1.second==59:
					ws_result['A%d' % result_i]=(time1-timedelta(minutes=i-1)).strftime('%Y-%m-%d %H:%M')
				else:
					ws_result['A%d' % result_i]=(time1-timedelta(minutes=i)).strftime('%Y-%m-%d %H:%M')
				result_i+=1
			count=1
			time1=ws.cell(row=index,column=1).value
	# 		result_i+=1
		index+=1
	
	ws_result['A%d' % result_i]=(time1).strftime('%Y-%m-%d %H:%M')
	for i in range(0,8):
		ws_result.cell(row=result_i,column=i+2).value=sum[i]/count
	# for i in range(0,8):
	# 	sum[i]=ws.cell(row=2,column=i+2).value
	# 	print sum[i]
		
		
	
	#print type(sum)
	# count=1
	# for i in range(2,7):
	# 	time2=ws.cell(row=i,column=1).value.strftime('%M')
	# 	if time2==time1:
	# 		if not ws.cell(row=i,column=2).value:
	# 			sum+=0
	# 			count+=1
	# 		else:
	# 			sum+=ws.cell(row=i,column=2).value
	# 			count+=1
	# 	else:
	# 		ws.cell(row=i,column=3).value=sum/count
	# 		time1=time2
	# 		sum=ws.cell(row=i,column=2).value
	# 		count=1
	# 		
	wb_result.save(filename_result)
	wb.save(filename)



