from __future__ import division
import sys
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import numpy as np
import pylab as pl
from scipy import interpolate 
import matplotlib.pyplot as plt

if not os.path.exists("result"):
    os.mkdir("result")
else:
    print "already have an result folder"
files = os.listdir(sys.path[0])

xnew=[]

for file in files:
	if '.xlsx' in file:
		print file
		wb=load_workbook(file)
		ws=wb.active
		
		i=1
		while ws['a%d' % i].value:
			if type(ws['a%d' % i].value) !=float:
				i=i+1
				continue
			xnew.append(ws['a%d' % i].value)
			i=i+1
		print "))))))))))))))))))))))))))))))"
	else: continue
	# print xnew[len(xnew)-1]
# print xnew
for file in files:
	# if file!='NO2_qy.txt':
	# 	continue
	x = []
	y = []
	ynew=[]
	test=[]
	if not '.txt' in file:
		continue
	# print file
	ftxt=open('result/'+file,'w')
	wbt=Workbook()
	wst=wbt.active
	lines = open(file).readlines()
	index=0
	print lines
	# print "lkdj"
	if xnew[index] < float(lines[0].split()[0]):
		print "this one need before hand"
		print file
	while xnew[index] < float(lines[0].split()[0]):
		# print "this one need before hand"
		# print file
		# print xnew[index]
		x.append(xnew[index])
		y.append(0)
		index+=1
	i=0
	while lines[i]!=" " and i<len(lines)-1:
		# print lines[i].split()[0]
		# if float(lines[i].split()[0])>xnew[0]:
			# print xnew[0]
			# print lines[i].split()[0]
		# 	continue
		# if float(lines[i].split()[0])>xnew[len(xnew)-1]:
		# 	# print lines[i].split()[0]
		# 	continue
		# print i
		x.append(float(lines[i].split()[0]))
		y.append(float(lines[i].split()[1]))
		i+=1


		# test.append([float(lines[i].split()[0]),float(lines[i].split()[1])])
	if x[len(x)-1]<xnew[len(xnew)-1]:
		print "after"
		print file
	while(x[len(x)-1]<xnew[len(xnew)-1]):

		x.append(x[len(x)-1]+1)
		y.append(0)

	# if file=='NO2_qy.txt':
	# 	print y
	f_linear = interpolate.interp1d(x, y)
	ynew=f_linear(xnew)
	# if file=='NO2_qy.txt':
	# 	for i in range(0,len(ynew)):
	# 		print ynew[i]
	for i in range(0,len(xnew)):
		test.append(str(xnew[i])+'\t'+str(ynew[i])+'\r\n')
	for r in test:
		ftxt.write(r)
	ftxt.close()
	# print test
	# for i in test:
	# 	# print i
	# 	wst.append(i)

	# wbt.save('result/'+file.split('.')[0]+'.xlsx')
	# wbt.close
# if ws['a8']:
# 	if type(ws['a8'].value)==float:
# 		print "is a lol;sdj"