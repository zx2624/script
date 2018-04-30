from __future__ import division 
from openpyxl import Workbook
from openpyxl import load_workbook

n=input("how manny?")
a=[0 for i in range(n)]
for i in range(0,n):
	d=i+1
	a[i]=input("input%d   " % d )
begin_all=input("begin index of data:  ")
end_all=input("end index of data:  ")
begin=input("begin index of stray:   ")
end=input("end index of stray:    ")

for file_in in range(0,n):
	filename="%d-1.xlsx" % a[file_in]
	filename_dark="%dms.xlsx" % a[file_in]
	filename_za="%d-2.xlsx" % a[file_in]
	filename_result="%d-result.xlsx" % a[file_in]
	#print filename_dark
	wb=load_workbook(filename)
	ws=wb.active
	wb_dark=load_workbook(filename_dark)
	ws_dark=wb_dark.active
	wb_result=Workbook()
	ws_result=wb_result.active
	wb_za=load_workbook(filename_za)
	ws_za=wb_za.active

	for i in range(begin_all,end_all+1):
		ws.cell(row=i,column=2).value-=ws_dark.cell(row=i,column=2).value

	aver_1_ms=0.0
	count_1_ms=0
	for i in range(begin,end+1):
		aver_1_ms=(count_1_ms*aver_1_ms+ws.cell(row=i,column=2).value)/(count_1_ms+1)
		count_1_ms+=1
	print "aver_1_ms is %f" % aver_1_ms
	
	for i in range(begin_all,end_all):
		ws_result.cell(row=i,column=2).value=ws.cell(row=i,column=2).value-aver_1_ms
		ws_result.cell(row=i,column=1).value=ws.cell(row=i,column=1).value
	wb_result.save(filename=filename_result)
	
	
	


	
