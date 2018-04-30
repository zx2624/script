from __future__ import division
import sys

from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import timedelta, datetime
import os
from __builtin__ import file
from pickle import NONE


def calcu_minutes(t1, t2):

    if t1.second == 59:
        m1 = t1.minute + 1
    else:
        m1 = t1.minute
    if t2.second == 59:
        m2 = t2.minute + 1
    else:
        m2 = t2.minute
    d = t1.day - t2.day
    h = t1.hour - t2.hour
    m = m1 - m2
    return h * 60 + m + d * 24 * 60
print "this may take a while......"
if not os.path.exists("result"):
    os.mkdir("result")
else:
    print "already have an result folder"

files = os.listdir(sys.path[0])
l_h2o = []
l_hchom = []
l_hchor = []
l_hono = []
l_no3m = []
l_no3r = []
l_no2 = []
l_o1d1 = []
lines_all = []


for file in files:
    if 'H2O' in file:
        l_h2o += open(file).readlines()
    if 'HCHO_M' in file:
        l_hchom += open(file).readlines()
    if 'HCHO_R' in file:
        l_hchor += open(file).readlines()
    if 'HONO' in file:
        l_hono += open(file).readlines()
    if 'NO3_M' in file:
        l_no3m += open(file).readlines()
    if 'NO3_R' in file:
        l_no3r += open(file).readlines()
    if 'NO2' in file:
        l_no2 += open(file).readlines()
    if 'O1D1' in file:
        l_o1d1 += open(file).readlines()

wb = Workbook()
ws = wb.active
wb_result = Workbook()
ws_result = wb_result.active
ws['a1'] = 0
ws['b1'] = 1

# A1=(C1-70*365-19)*86400-8*3600
for i in range(0, len(l_h2o)):
    lines_all.append([datetime.fromtimestamp(((float(l_h2o[i].split()[0]) - 70 * 365 - 19) * 86400 - 8 * 3600)), float(l_o1d1[i].split()[1]), float(l_hchom[i].split()[1]), float(
        l_no2[i].split()[1]), float(l_h2o[i].split()[1]), float(l_hono[i].split()[1]), float(l_no3m[i].split()[1]), float(l_hchor[i].split()[1]), float(l_no3r[i].split()[1])])

for r in lines_all:
    ws.append(r)

time1 = ws['A2'].value
sum = [0, 0, 0, 0, 0, 0, 0, 0]
for i in range(0, 8):
    sum[i] = ws.cell(row=2, column=i + 2).value
#    print sum
count = 1
index = 3
result_i = 2
btou = ['', 'JO1D', 'JNO2', 'JH2O2', 'JHONO', 'JNO3_M', 'JHCHO_R', 'JNO3_R']
ws_result.append(btou)
# print calcu_minutes(time1, ws['A5'].value)
# print ws['A5'].value.minute
# print ws['A5'].value.second
while(ws['A%d' % index].value):
    if calcu_minutes(ws['A%d' % index].value, time1) == 0:
        for i in range(0, 8):
            sum[i] += ws.cell(row=index, column=i + 2).value
        # print sum
        count += 1
#         print count
    if calcu_minutes(ws['A%d' % index].value, time1) == 1:
        #             if time1.second==59:
        # #                 time1=time1+timedelta(minutes=1)
        # #                 time1=time1-timedelta(seconds=time1.second+1)
        #                 ws_result['A%d' % result_i]=time1
        #             else:
        # #                 time1=time1-timedelta(seconds=time1.second+1)
        ws_result['A%d' % result_i] = time1
        for i in range(0, 8):
            ws_result.cell(row=result_i, column=i + 2).value = sum[i] / count
        for i in range(0, 8):
            sum[i] = ws.cell(row=index, column=i + 2).value
        count = 1
        time1 = ws.cell(row=index, column=1).value
        result_i += 1
    if calcu_minutes(ws['A%d' % index].value, time1) > 1:
        #         print calcu_minutes(time1,ws['A%d' % index].value)
        #         print time1
        #             if time1.second==59:
        # #                 time1=time1+timedelta(minutes=1)
        # #                 time1=time1-timedelta(seconds=time1.second+1)
        #                 ws_result['A%d' % result_i]=time1
        #             else:
        # #                 time1=time1-timedelta(seconds=time1.second+1)
        ws_result['A%d' % result_i] = time1
        for i in range(0, 8):
            ws_result.cell(row=result_i, column=i + 2).value = sum[i] / count
        for i in range(0, 8):
            sum[i] = ws.cell(row=index, column=i + 2).value
        result_i += 1
        for i in range(1, calcu_minutes(ws['A%d' % index].value, time1)):
            #                 if time1.second==59:
            #                     time1=time1-timedelta(minutes=i)
            # #                     time1=time1-timedelta(seconds=time1.second+1)
            #                     ws_result['A%d' % result_i]=time1
            #                 else:
            time1 = time1 + timedelta(minutes=1)
            print time1
            ws_result['A%d' % result_i] = time1
            result_i += 1
        count = 1
        time1 = ws.cell(row=index, column=1).value

    index += 1


wb_result.save('result/result.xlsx')
wb.save('result/combine.xlsx')
# print line
