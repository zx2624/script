from __future__ import division 
from openpyxl import Workbook
from openpyxl import load_workbook

########-1.xlsx-ms.xlsx#then get the average between 80~99#deduct this part from #-1.xlsx-ms.xlsx# ##get the smooth average##finally get the 270.280.290....

#input the necessary params,
n=input("how manny?")
a=[0 for i in range(n)]
for i in range(0,n):
	d=i+1
	a[i]=input("file %d   " % d )
begin_all=input("begin index of data:  ")
end_all=input("end index of data:  ")
begin=input("begin index of stray:   ")
end=input("end index of stray:    ")

#deal with every file
for file_in in range(0,n):
	filename="%d-1.xlsx" % a[file_in]
	filename_dark="%dms.xlsx" % a[file_in]
# 	filename_za="%d-2.xlsx" % a[file_in]
	filename_result="%d-average.xlsx" % a[file_in]
	filename_10="%d-10.xlsx" % a[file_in]
	#print filename_dark
	wb=load_workbook(filename)
	ws=wb.active
	wb_dark=load_workbook(filename_dark)
	ws_dark=wb_dark.active
	wb_result=Workbook()
	ws_result=wb_result.active
	wb_tmp_average=Workbook()
	ws_tmp_average=wb_tmp_average.active
	wb_10=Workbook()
	ws_10=wb_10.active
#	wb_za=load_workbook(filename_za)
#	ws_za=wb_za.active
#
#deduct the ms part
	for i in range(begin_all,end_all+1):
		ws.cell(row=i,column=2).value-=ws_dark.cell(row=i,column=2).value
#calculate average
	aver_1_ms=0.0
	count_1_ms=0
# 	calculatie the average between 85-90nm or whatever
	for i in range(begin,end+1):
		aver_1_ms=(count_1_ms*aver_1_ms+ws.cell(row=i,column=2).value)/(count_1_ms+1)
		count_1_ms+=1
	print "aver_1_ms is %f" % aver_1_ms
#deduct average	
	for i in range(begin_all,end_all):
		ws_tmp_average.cell(row=i,column=2).value=ws.cell(row=i,column=2).value-aver_1_ms
		ws_tmp_average.cell(row=i,column=1).value=ws.cell(row=i,column=1).value
		

#calculate the smooth average and pick out 270/280/290/300......
	index1=int(ws_tmp_average.cell(row=begin_all,column=1).value)
	index2=index1+1
	count_1=0
	count_2=0
	aver_1=0
	aver_2=0
	result_index=0	
	count_za=0
	aver_za=0.0

	wavelength=270
	i_wave=1
	
	for i in range(begin_all,end_all):
		lamda=int(ws_tmp_average.cell(row=i,column=1).value)
	
		if lamda==index1:
			d=aver_1*count_1
		
			aver_1=(d+ws_tmp_average.cell(row=i,column=2).value)/(count_1+1)
			  
			count_1+=1
		if lamda==index2:
			
			aver_2=(aver_2*count_2+ws_tmp_average.cell(row=i,column=2).value)/(count_2+1)
			count_2+=1
		if lamda!=index1 and lamda!=index2:
			if count_1+count_2==0:continue
			result_index+=1
			ws_result.cell(row=result_index,column=1).value=lamda-2
			ws_result.cell(row=result_index,column=2).value=(aver_1*count_1+aver_2*count_2)/(count_1+count_2)
			if lamda-2==wavelength:
				ws_10.cell(row=i_wave,column=1).value=wavelength
				ws_10.cell(row=i_wave,column=2).value=ws_result.cell(row=result_index,column=2).value
				i_wave+=1
				wavelength+=10
		 	aver_1=aver_2
			aver_2=ws_tmp_average.cell(row=i,column=2).value
			count_1=count_2
			
			count_2=1
			index1+=1
			index2+=1
	
	result_index+=1
	ws_result.cell(row=result_index,column=1).value=lamda-1
	ws_result.cell(row=result_index,column=2).value=(aver_1*count_1+aver_2*count_2)/(count_1+count_2)
	
	wb_result.save(filename=filename_result)
	wb_10.save(filename=filename_10)
#	wb_tmp_average.save("tmp.xlsx")
	
	
	


	
