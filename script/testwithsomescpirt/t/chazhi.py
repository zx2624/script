from __future__ import division 
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import sys

wb_jiaozhun=load_workbook("jiaozhun2.xlsx")
ws_jiaozhun=wb_jiaozhun.active


#a=[0 for i in range(7)]
a=[20,40,80,200,500,800,1000]
for i in range(0,7):
    filenamesave="%d-result.xlsx" % a[i]
    wb_20_result=load_workbook(filenamesave)
    ws_20_result=wb_20_result.active
    bar_now=270
    bar_next=280
    index_jiao=2
    resolution=(ws_jiaozhun.cell(row=index_jiao+1,column=2).value-ws_jiaozhun.cell(row=index_jiao,column=2).value)/10
    for i in range(7,2055):
        if index_jiao==45:
            ws_20_result.cell(row =i,column=3).value=0
            ws_20_result.cell(row =i,column=4).value=0
	    ws_20_result.cell(row =i,column=5).value=0
            continue
        if ws_20_result.cell(row =i,column=1).value<bar_now:
            ws_20_result.cell(row =i,column=3).value=0
            ws_20_result.cell(row =i,column=4).value=0
	    ws_20_result.cell(row =i,column=5).value=0
        elif ws_20_result.cell(row =i,column=1).value<bar_next:
            ws_20_result.cell(row =i,column=3).value=resolution*(ws_20_result.cell(row =i,column=1).value-bar_now)
	    ws_20_result.cell(row =i,column=4).value=ws_20_result.cell(row =i,column=3).value+ws_jiaozhun.cell(row=index_jiao,column=2).value
            ws_20_result.cell(row =i,column=5).value=ws_20_result.cell(row =i,column=4).value/ws_20_result.cell(row =i,column=2).value
        else :
            index_jiao+=1
            bar_now=bar_next
            bar_next+=10
            if index_jiao==45:
                ws_20_result.cell(row =i,column=3).value=0
                ws_20_result.cell(row =i,column=4).value=0
		ws_20_result.cell(row =i,column=5).value=0
                continue
            resolution=(ws_jiaozhun.cell(row=index_jiao+1,column=2).value-ws_jiaozhun.cell(row=index_jiao,column=2).value)/10
            ws_20_result.cell(row =i,column=3).value=resolution*(ws_20_result.cell(row =i,column=1).value-bar_now)
            ws_20_result.cell(row =i,column=4).value=ws_20_result.cell(row =i,column=3).value+ws_jiaozhun.cell(row=index_jiao,column=2).value
	    ws_20_result.cell(row =i,column=5).value=ws_20_result.cell(row =i,column=4).value/ws_20_result.cell(row =i,column=2).value



    wb_20_result.save(filename=filenamesave)






	
