from __future__ import division 
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import sys

wb_jiaozhun=load_workbook("jiaozhun2.xlsx")
ws_jiaozhun=wb_jiaozhun.active

if not os.path.exists("result"):
	os.mkdir("result")
files=os.listdir(sys.path[0])

begin_all=7
end_all=2054


for file in files:
	s=file.split("-")
	if len(s)>1 and s[1]=="1.xlsx":
		filename=file
		filename_dark=s[0]+"ms.xlsx"
		print filename_dark
		filename_result="result/"+s[0]+"-result.xlsx"
	else:continue
	
#	filename_za="%d-2.xlsx" % a[file_in]
		#print filename_dark
	wb=load_workbook(filename)
	ws=wb.active
	wb_dark=load_workbook(filename_dark)
	ws_dark=wb_dark.active
	wb_result=Workbook()
	ws_result=wb_result.active
#	wb_za=load_workbook(filename_za)
#	ws_za=wb_za.active
	bar_now=270
	bar_next=280
	index_jiao=2
	resolution=(ws_jiaozhun.cell(row=index_jiao+1,column=2).value-ws_jiaozhun.cell(row=index_jiao,column=2).value)/10
	ws_result.cell(row=6,column=4).value="interploration"
	ws_result.cell(row=6,column=6).value="final_result"
	for i in range(begin_all,end_all+1):
		ws_result.cell(row=i,column=2).value=ws.cell(row=i,column=2).value-ws_dark.cell(row=i,column=2).value
		ws_result.cell(row=i,column=1).value=ws.cell(row=i,column=1).value
 
# 	aver_1_ms=0.0
# 	count_1_ms=0
# 	for i in range(begin,end+1):
# 		aver_1_ms=(count_1_ms*aver_1_ms+ws.cell(row=i,column=2).value)/(count_1_ms+1)
# 		count_1_ms+=1
# 	print "aver_1_ms is %f" % aver_1_ms
 	
# 	for i in range(begin_all,end_all+1):
# 		ws_result.cell(row=i,column=2).value=ws.cell(row=i,column=2).value-aver_1_ms
# 		ws_result.cell(row=i,column=1).value=ws.cell(row=i,column=1).value
 		
	for i in range(begin_all,end_all+1):
	    if index_jiao==45:
 	        
	        ws_result.cell(row =i,column=4).value=0
	        ws_result.cell(row =i,column=5).value=0
	        ws_result.cell(row =i,column=6).value=0
	        continue
	    if ws_result.cell(row =i,column=1).value<bar_now:
 	        
	        ws_result.cell(row =i,column=4).value=0
	        ws_result.cell(row =i,column=5).value=0
	        ws_result.cell(row =i,column=6).value=0
	    elif ws_result.cell(row =i,column=1).value<bar_next:
 	        
	        ws_result.cell(row =i,column=4).value=resolution*(ws_result.cell(row =i,column=1).value-bar_now)+ws_jiaozhun.cell(row=index_jiao,column=2).value
	        ws_result.cell(row =i,column=5).value=ws_result.cell(row =i,column=4).value/ws_result.cell(row =i,column=2).value
	        ws_result.cell(row =i,column=6).value=ws_result.cell(row=i,column=1).value*ws_result.cell(row =i,column=5).value*1e11/6.626/3
	    else :
	        index_jiao+=1
	        bar_now=bar_next
	        bar_next+=10
	        if index_jiao==45:
	            ws_result.cell(row =i,column=6).value=0
	            ws_result.cell(row =i,column=4).value=0
	            ws_result.cell(row =i,column=5).value=0
	            continue
	        resolution=(ws_jiaozhun.cell(row=index_jiao+1,column=2).value-ws_jiaozhun.cell(row=index_jiao,column=2).value)/10
	#            ws_result.cell(row =i,column=3).value=resolution*(ws_result.cell(row =i,column=1).value-bar_now)
	        ws_result.cell(row =i,column=4).value=resolution*(ws_result.cell(row =i,column=1).value-bar_now)+ws_jiaozhun.cell(row=index_jiao,column=2).value
	        ws_result.cell(row =i,column=5).value=ws_result.cell(row =i,column=4).value/ws_result.cell(row =i,column=2).value
	        ws_result.cell(row =i,column=6).value=ws_result.cell(row=i,column=1).value*ws_result.cell(row =i,column=5).value*1e11/6.626/3
     
     
	wb_result.save(filename=filename_result)
	
	
	


	
