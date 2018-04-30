from __future__ import division
import sys
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import numpy as np
import pylab as pl
from scipy import interpolate 
import matplotlib.pyplot as plt

if not os.path.exists("result"):
    os.mkdir("result")
else:
    print "already have an result folder"
files = os.listdir(sys.path[0])
x = []
y = []
xnew=[]
test=[]
for file in files:
	if '.xlsx' in file:
		print file
		wb=load_workbook(file)
		ws=wb.active
		
		i=1
		while ws['a%d' % i].value:
			if type(ws['a%d' % i].value) !=float:
				i=i+1
				continue
			xnew.append(ws['a%d' % i].value)
			i=i+1
	# print xnew[len(xnew)-1]
for file in files:
	if not '.txt' in file:
		continue
	print file
	wbt=Workbook()
	ws=wbt.active
	lines = open(file).readlines()
	index=0
	while xnew[index] < float(lines[i].split()[0]):
		x.append(xnew[index])
		y.append(0)
		index+=1
	for i in range(0, len(lines)):
		# print lines[i].split()[0]
		# if float(lines[i].split()[0])>xnew[0]:
			# print xnew[0]
			# print lines[i].split()[0]
		# 	continue
		# if float(lines[i].split()[0])>xnew[len(xnew)-1]:
		# 	# print lines[i].split()[0]
		# 	continue
		x.append(float(lines[i].split()[0]))
		y.append(float(lines[i].split()[1]))
		# test.append([float(lines[i].split()[0]),float(lines[i].split()[1])])
	while(x[len(x)-1]<xnew[len(xnew)-1]):
		x.append(x[len(x)-1]+1)
		y.append(0)


	f_linear = interpolate.interp1d(x, y)
	ynew=f_linear(xnew)
	for i in range(0,len(xnew)):
		test.append([xnew[i],ynew[i]])
	for i in test:
		# print i
		ws.append(i)

	wbt.save('result/'+file.split('.')[0]+'.xlsx')
# if ws['a8']:
# 	if type(ws['a8'].value)==float:
# 		print "is a lol;sdj"