from __future__ import division
import sys
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import numpy as np
import pylab as pl
from scipy import interpolate 
import matplotlib.pyplot as plt

if not os.path.exists("result"):
    os.mkdir("result")
else:
    print "already have an result folder"
files = os.listdir(sys.path[0])

x=[]
y=[]
#xx stores the x coordinate
xx=[]

count=0
for file in files:
	if 'jiaozhun' in file:
		wbcali=load_workbook(file)
		wscali=wbcali.active
		i=1
		while wscali['a%d' % i].value:
			if type(wscali['a%d' % i].value) !=long:
				i=i+1
				continue
			x.append(wscali['a%d' % i].value)
			y.append(wscali['b%d' % i].value)
			i=i+1
		break
for file in files:
	if '-1' in file:
		print file
		if count==1:
			continue

		wbx=load_workbook(file)
		wsx=wbx.active
		i=1
		print wsx['a10'].value
		print type(wsx['a9'].value)
		while wsx['a%d' % i].value:
			if type(wsx['a%d' % i].value) !=float:
				i=i+1
				continue
			xx.append(wsx['a%d' % i].value)
			i=i+1
		break
f_linear = interpolate.interp1d(x, y,bounds_error=False,fill_value=0)
f_cubic = interpolate.interp1d(x,y,kind='cubic',bounds_error=False,fill_value=0)
print len(xx)
yy=f_linear(xx)
yyc=f_cubic(xx)
rlist=[]
for file in files:
	if '-1' in file:
		s=file.split("-")
		wb1=load_workbook(file)
		ws1=wb1.active
		print s[0]+'-2.xlsx'
		wb2=load_workbook(s[0]+'-2.xlsx')
		ws2=wb2.active
		wbms=load_workbook(s[0]+'ms.xlsx')
		wsms=wbms.active
		wbresult=Workbook()
		wsresult=wbresult.active
		i=1
		while ws1['a%d' % i].value:
			if type(ws1['a%d' % i].value) !=float:
				i=i+1
				continue
			v1=ws1['b%d' % i].value
			v2=ws2['b%d' % i].value
			vms=wsms['b%d' % i].value
			if ws1['b%d' % i].value<350:
				rlist.append([v1-v2,yy[i-7],yyc[i-7]])
			else:
				rlist.append([v1-vms,yy[i-7],yyc[i-7]])
			i=i+1
		for t in rlist:
			wsresult.append(t)
		wbresult.save('result/'+s[0]+'result.xlsx')
